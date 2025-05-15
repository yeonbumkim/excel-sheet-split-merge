import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from io import BytesIO
from datetime import datetime
from typing import List, Tuple, Dict, Optional
import re

# =========================
# 유틸 함수
# =========================
def sanitize_sheet_name(name: str) -> str:
    # Excel 시트명에서 허용되지 않는 문자 제거
    return re.sub(r'[\\/*?:\[\]]', '_', name)

def get_today(fmt: str) -> str:
    if fmt == 'YYYY-MM-DD':
        return datetime.now().strftime('%Y-%m-%d')
    return datetime.now().strftime('%Y%m%d')

# =========================
# 시트 분할 함수
# =========================
def split_sheets(
    file: BytesIO,
    filename: str,
    date_fmt: str,
    naming_rule: str,
) -> List[Tuple[str, BytesIO]]:
    """
    엑셀 파일의 각 시트를 분할하여 지정된 규칙으로 파일명과 함께 반환
    """
    wb = load_workbook(file, data_only=True)
    today = get_today(date_fmt)
    results = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = ws.title
        for row in ws:
            for cell in row:
                new_ws[cell.coordinate].value = cell.value
        # 파일명 규칙 적용
        sanitized_sheet = sanitize_sheet_name(sheet_name)
        sanitized_file = filename.rsplit('.', 1)[0]
        if naming_rule == 'a':
            out_name = f"{sanitized_file}_{sanitized_sheet}.xlsx"
        elif naming_rule == 'b':
            out_name = f"{sanitized_sheet}.xlsx"
        elif naming_rule == 'c':
            out_name = f"{today}_{sanitized_file}_{sanitized_sheet}.xlsx"
        else:
            out_name = f"{today}_{sanitized_sheet}.xlsx"
        out_io = BytesIO()
        new_wb.save(out_io)
        out_io.seek(0)
        results.append((out_name, out_io))
    return results

# =========================
# 워크북 통합 함수
# =========================
def merge_workbooks(
    files: List[Tuple[str, BytesIO]],
    merged_filename: str,
    date_fmt: str,
) -> Tuple[str, BytesIO]:
    """
    여러 엑셀 파일의 시트를 하나의 워크북으로 통합
    """
    today = get_today(date_fmt)
    merged_wb = Workbook()
    merged_wb.remove(merged_wb.active)  # 기본 시트 제거
    sheet_names = set()
    for fname, file in files:
        wb = load_workbook(file, data_only=True)
        base_fname = fname.rsplit('.', 1)[0]
        for sheet_name in wb.sheetnames:
            orig_sheet = wb[sheet_name]
            # 시트명 충돌 방지
            sanitized_sheet = sanitize_sheet_name(sheet_name)
            new_sheet_name = sanitized_sheet
            if new_sheet_name in sheet_names:
                new_sheet_name = f"{base_fname}_{sanitized_sheet}"
            count = 1
            temp_name = new_sheet_name
            while temp_name in sheet_names:
                temp_name = f"{new_sheet_name}_{count}"
                count += 1
            new_sheet_name = temp_name
            sheet_names.add(new_sheet_name)
            new_ws = merged_wb.create_sheet(new_sheet_name)
            for row in orig_sheet:
                for cell in row:
                    new_ws[cell.coordinate].value = cell.value
    # 파일명에 날짜 적용
    if date_fmt:
        merged_filename = f"{get_today(date_fmt)}_{merged_filename}"
    if not merged_filename.endswith('.xlsx'):
        merged_filename += '.xlsx'
    out_io = BytesIO()
    merged_wb.save(out_io)
    out_io.seek(0)
    return merged_filename, out_io

# =========================
# Streamlit UI
# =========================
st.set_page_config(
    page_title="엑셀 시트 분할 및 통합 도구",
    layout="centered",
    initial_sidebar_state="expanded",
)
st.title("Excel Sheet Split & Merge")

# 도움말 버튼 및 안내
if st.button("도움말 보기 🛈"):
    st.info(
        """
        ### 사용 시 주의사항 및 안내
        
        **1. 메모리 사용량**  
        업로드한 파일과 분할/통합된 결과 파일은 모두 메모리(BytesIO)에서 처리됩니다.  
        매우 큰 엑셀 파일(수십 MB 이상)이나 시트가 많은 경우, 서버 메모리 부족 현상이 발생할 수 있습니다.
        
        **2. 파일 확장자 및 형식**  
        `.xlsx` 파일만 지원합니다.  
        `.xls`(구버전), CSV 등은 지원하지 않습니다.  
        파일 확장자만 검사하므로, 실제로는 비정상 파일이 업로드될 경우 에러가 발생할 수 있습니다.
        
        **3. 시트명/파일명 규칙**  
        시트명에 Excel에서 허용하지 않는 문자가 있으면 자동으로 `_`로 치환합니다.  
        통합 시 시트명이 충돌하면 `파일명_시트명`, 그래도 충돌하면 `파일명_시트명_숫자` 형식으로 자동 변경됩니다.
        
        **4. 예외 처리**  
        빈 파일, 시트가 없는 파일, 잘못된 파일 등은 Streamlit의 경고/오류 메시지로 안내됩니다.  
        예상치 못한 에러 발생 시에도 사용자에게 오류 메시지가 표시됩니다.
        
        **5. 동시성**  
        Streamlit은 기본적으로 각 사용자의 세션을 분리하여 처리합니다.  
        여러 사용자가 동시에 사용할 경우, 서로의 데이터가 섞이지 않습니다.
        
        **6. 보안 및 개인정보**  
        업로드된 파일은 서버 메모리에서만 처리되고, 별도로 저장되지 않습니다.  
        작업 수행 후 서버에 파일이 남지 않아 개인정보, 민감 데이터 유출 위험이 적습니다.
        
        ---
        
        **요약**  
        - 업로드/다운로드 파일은 서버에 저장되지 않고, 메모리에서만 처리됩니다.  
        - 대용량 파일 처리 시 메모리 부족에 주의하세요.  
        - `.xlsx`만 지원, 파일명/시트명 규칙에 따라 이름이 자동 변경될 수 있습니다.  
        - 예외 상황은 Streamlit 메시지로 안내됩니다.
        """
    )

# 사이드바
st.sidebar.header("설정")
mode = st.sidebar.radio("기능 선택", ["시트 분할", "파일 통합"])

if mode == "시트 분할":
    uploaded_file = st.sidebar.file_uploader("엑셀 파일 업로드 (.xlsx)", type=["xlsx"])
    naming_rule = st.sidebar.selectbox(
        "파일명 규칙",
        [
            "[원본파일명]_[Tab이름].xlsx",
            "[Tab이름].xlsx",
            "[YYYYMMDD]_[원본파일명]_[Tab이름].xlsx",
            "[YYYYMMDD]_[Tab이름].xlsx",
        ],
        index=0,
    )
    st.subheader("시트 분할 실행")
    if uploaded_file:
        if st.button("문서 분할 시작", key="split_btn"):
            try:
                if not uploaded_file.name.endswith('.xlsx'):
                    st.error("xlsx 파일만 업로드 가능합니다.")
                else:
                    rule_idx = [
                        "[원본파일명]_[Tab이름].xlsx",
                        "[Tab이름].xlsx",
                        "[YYYYMMDD]_[원본파일명]_[Tab이름].xlsx",
                        "[YYYYMMDD]_[Tab이름].xlsx",
                    ].index(naming_rule)
                    rule_map = {0: 'a', 1: 'b', 2: 'c', 3: 'd'}
                    # 날짜포맷이 필요한 경우만 YYYYMMDD로 넘김
                    date_fmt = 'YYYYMMDD' if rule_idx in [2, 3] else ''
                    with st.spinner('시트 분할 작업 중입니다...'):
                        results = split_sheets(
                            file=uploaded_file,
                            filename=uploaded_file.name,
                            date_fmt=date_fmt,
                            naming_rule=rule_map[rule_idx],
                        )
                    if not results:
                        st.warning("시트가 없는 파일입니다.")
                    else:
                        st.success(f"{len(results)}개 시트 분할 완료!")
                        for fname, io in results:
                            st.download_button(
                                label=f"{fname} 다운로드",
                                data=io,
                                file_name=fname,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            )
            except Exception as e:
                st.error(f"오류 발생: {e}")
        else:
            st.info("문서 분할 시작 버튼을 눌러주세요.")
    else:
        st.info("엑셀 파일을 사이드바에서 업로드하세요.")

elif mode == "파일 통합":
    uploaded_files = st.sidebar.file_uploader(
        "엑셀 파일 여러 개 업로드 (.xlsx, 2개 이상)", type=["xlsx"], accept_multiple_files=True
    )
    merged_filename = st.sidebar.text_input("통합 파일명", value="merged.xlsx")
    st.subheader("문서 통합 실행")
    if uploaded_files and len(uploaded_files) >= 2:
        if st.button("문서 통합 시작", key="merge_btn"):
            try:
                files = [(f.name, f) for f in uploaded_files]
                # 통합 파일명에 날짜포맷이 포함된 경우만 YYYYMMDD 적용
                date_fmt = 'YYYYMMDD' if merged_filename.startswith('YYYYMMDD') or merged_filename.startswith('[YYYYMMDD]') else ''
                with st.spinner('파일 통합 작업 중입니다...'):
                    out_name, out_io = merge_workbooks(
                        files=files,
                        merged_filename=merged_filename,
                        date_fmt=date_fmt,
                    )
                st.success(f"{out_name} 통합 완료!")
                st.download_button(
                    label=f"{out_name} 다운로드",
                    data=out_io,
                    file_name=out_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as e:
                st.error(f"오류 발생: {e}")
        else:
            st.info("문서 통합 시작 버튼을 눌러주세요.")
    else:
        st.info("엑셀 파일을 2개 이상 사이드바에서 업로드하세요.")

# --- 개발자 정보 하단 고정 ---
st.markdown("""
---
<div style='text-align:center; color:gray; font-size:0.95em;'>
  <b>개발자:</b> Yeonbum Kim &nbsp;|&nbsp; <a href='mailto:yeonbumk@gmail.com'>yeonbumk@gmail.com</a>
</div>
""", unsafe_allow_html=True) 